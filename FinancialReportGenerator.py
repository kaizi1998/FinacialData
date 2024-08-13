import pandas as pd
from FinancialReportGetter import FinancialReportGetter
from FinancialReportGetterAfterStartDate import FinancialReportGetterAfterStartDate
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment, NamedStyle
# 先假设只有一个城市长沙
from openpyxl import Workbook
from datetime import date, timedelta
from ConfigReader import station_map


class Iterator:
     def __init__(self):
          self._ch = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
          self._cur_pos = 0

     @property
     def ch(self):
          r = self._ch[self._cur_pos]
          self._cur_pos += 1
          return r


# 设置自动换行,'D': 9.96296296296296
class FinancialReportGenerator:
     df_station_map = station_map()

     @classmethod
     def _generate_form(cls, ws):
          # 设置的列宽值
          it = Iterator()
          column_widths = {
               it.ch: 6.66666666666667, it.ch: 20.7685185185185, it.ch:16, it.ch:16, it.ch: 14,
               it.ch: 10.4074074074074, it.ch: 11.75, it.ch: 10.0925925925926, it.ch: 12.2962962962963,
               it.ch: 11.75, it.ch: 9.51851851851852, it.ch: 10.2962962962963, it.ch: 11.962962962963,
               it.ch: 13.9537037037037, it.ch: 12.0925925925926, it.ch: 13.962962962963,
               it.ch: 15.6203703703704, it.ch: 10.75, it.ch: 12.7777777777778, it.ch: 17.5,
               it.ch: 12.0925925925926, it.ch: 14.9351851851852, it.ch: 14.7314814814815,
               it.ch: 12.3888888888889
          }

          # 遍历列宽字典，设置每列的宽度
          for col_letter, width in column_widths.items():
               ws.column_dimensions[col_letter].width = width

     @classmethod
     def _set_cell_border(cls, cell, border_left="thin", border_right="thin", border_top="thin", border_bottom="thin"):
          cell.border = Border(left=Side(style=border_left), right=Side(style=border_right),
                                        top=Side(style=border_top), bottom=Side(style=border_bottom))

     @classmethod
     def _set_cell(cls, cell, value, horizontal="center", vertical="center", wrapText=True,
                   font_name="宋体", font_size=11, font_bold=True, font_color="000000",
                   fill_color='FFFFFF',
                   border_left="thin", border_right="thin", border_top="thin", border_bottom="thin",
                   num_style='#,##0.00'):
          cell.value = value
          cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrapText=wrapText)
          cell.font = Font(name=font_name, size=font_size, bold=font_bold, color=font_color)
          cell.fill = PatternFill(start_color=fill_color,  fill_type='solid')
          cell.number_format = num_style
          # 创建一个中等虚线的边框样式
          cls._set_cell_border(cell, border_left, border_right, border_top, border_bottom)

     @classmethod
     def _generate_city(cls, ws, df:pd.DataFrame, city, date_str, cur_row):
          city = str(city)

          """  
          第1行, 行高30.0：
          A - V：
          内容："**汇天充电站运营情况信息表（YYYY.mm.dd）", "**"为城市名，"YYYY.mm.dd"为输入的年月日，
          字体：22号，加粗
          """
          it = Iterator()
          ch_start = it.ch
          for i in range(22):
               ch_end = it.ch
          ws.merge_cells(f"{ch_start}{cur_row}:{ch_end}{cur_row}")
          ws.row_dimensions[cur_row].height = 30.0
          cell = ws[f"{ch_start}{cur_row}"]
          cls._set_cell(cell=cell, value=f"{city}汇天充电站运营情况信息表（{date_str}）",font_size=22, border_left="none",border_right="none")
          cur_row += 1
          ws.row_dimensions[cur_row].height = None
          ws.row_dimensions[cur_row+1].height = 43.2

          """
          第2行，行高None：
              字体：11号，加粗
              A: 操作：和第三行合并；内容："序号"
              B: 操作：和第三行合并；内容："场站"
              C: 操作：和第三行合并；内容："总装机\nkw"
              D-J： 内容："每日充电数据"；背景颜色："00B050"（绿色）
              K-P： 内容："月度充电数据"；背景颜色："FFFF00"（黄色）
              Q-V： 内容："年度充电数据"；背景颜色："00B0F0"（蓝色）
          """
          it = Iterator()
          ch = it.ch
          ws.merge_cells(f"{ch}{cur_row}:{ch}{cur_row+1}")
          cls._set_cell(cell=ws[f"{ch}{cur_row}"], value="序号")
          cls._set_cell_border(cell=ws[f"{ch}{cur_row+1}"])
          ch = it.ch
          ws.merge_cells(f"{ch}{cur_row}:{ch}{cur_row + 1}")
          cls._set_cell(cell=ws[f"{ch}{cur_row}"], value="场站")
          cls._set_cell_border(cell=ws[f"{ch}{cur_row + 1}"])
          ch = it.ch
          ws.merge_cells(f"{ch}{cur_row}:{ch}{cur_row + 1}")
          cls._set_cell(cell=ws[f"{ch}{cur_row}"], value="运营方")
          cls._set_cell_border(cell=ws[f"{ch}{cur_row + 1}"])
          ch = it.ch
          ws.merge_cells(f"{ch}{cur_row}:{ch}{cur_row + 1}")
          cls._set_cell(cell=ws[f"{ch}{cur_row}"], value="上线时间")
          cls._set_cell_border(cell=ws[f"{ch}{cur_row + 1}"])
          ch = it.ch
          ws.merge_cells(f"{ch}{cur_row}:{ch}{cur_row + 1}")
          cls._set_cell(cell=ws[f"{ch}{cur_row}"], value="总装机\nkw")
          cls._set_cell_border(cell=ws[f"{ch}{cur_row + 1}"])
          ch_l = []
          for i in range(7):
               ch_l.append(it.ch)
          ws.merge_cells(f"{ch_l[0]}{cur_row}:{ch_l[-1]}{cur_row}")
          cls._set_cell(cell=ws[f"{ch_l[0]}{cur_row}"], value="每日充电数据", fill_color="00B050")
          for s in ch_l[1:]:
               cls._set_cell_border(cell=ws[f"{s}{cur_row}"])
          ch_l = []
          for i in range(6):
               ch_l.append(it.ch)
          ws.merge_cells(f"{ch_l[0]}{cur_row}:{ch_l[-1]}{cur_row}")
          cls._set_cell(cell=ws[f"{ch_l[0]}{cur_row}"], value="月度充电数据", fill_color="FFFF00")
          for s in ch_l[1:]:
               cls._set_cell_border(cell=ws[f"{s}{cur_row}"])
          ch_l = []
          for i in range(6):
               ch_l.append(it.ch)
          ws.merge_cells(f"{ch_l[0]}{cur_row}:{ch_l[-1]}{cur_row}")
          cls._set_cell(cell=ws[f"{ch_l[0]}{cur_row}"], value="年度充电数据", fill_color="00B0F0")
          for s in ch_l[1:]:
               cls._set_cell_border(cell=ws[f"{s}{cur_row}"])
          cur_row += 1

          """
          第3行，行高43.2：
              字体：11号，加粗
              D-J：
                  内容：D,"充电次数";E,"充电量\n(kWh)";F,"当日利用小时\n（h）";
                      G,"电费金额\n（元）";H,"服务费金额\n（元）";
                      I,"度均服务费（元）";J,"停车费收入（元）"
                  背景颜色："00B050"（绿色）
              K-P：
                  内容：K,"充电次数";L,"充电量\n(kWh)";M,"当日利用小时\n（h）";
                      N,"电费金额\n（元）";O,"服务费金额\n（元）";P,"停车费收入（元）"
                  背景颜色："FFFF00"（黄色）
              Q-V：
                  内容：Q,"充电次数";R,"充电量\n(kWh)";S,"当日利用小时\n（h）";
                      T,"电费金额\n（元）";U,"服务费金额\n（元）";V,"停车费收入（元）"
                  背景颜色："00B0F0"（蓝色）
          """
          it = Iterator()
          for i in range(5):
               it.ch
          cur_color = "00B050"
          cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value="充电次数", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value="充电量\n(kWh)", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value="当日利用小时\n（h）", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value="电费金额\n（元）", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value="服务费金额\n（元）", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value="度均服务费（元）", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value="停车费收入（元）", fill_color=cur_color)

          cur_color = "FFFF00"
          cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value="充电次数", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value="充电量\n(kWh)", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value="累计利用小时\n（h）", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value="电费金额\n（元）", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value="服务费金额\n（元）", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value="停车费收入（元）", fill_color=cur_color)

          cur_color = "00B0F0"
          cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value="充电次数", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value="充电量\n(kWh)", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value="累计利用小时\n（h）", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value="电费金额\n（元）", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value="服务费金额\n（元）", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value="停车费收入（元）", fill_color=cur_color)
          cur_row += 1

          """
          中间的行，行高27.0：
              字体：11号，加粗
              背景颜色：D-J，"00B050"（绿色）；K-P，"FFFF00"（黄色）；Q-V，"00B0F0"（蓝色）
          """
          middle_rows = len(df)
          middle_start = cur_row
          for i in range(middle_rows):
               it = Iterator()
               ws.row_dimensions[cur_row].height = 27.0
               cur_color = "FFFFFF"
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=f"{i + 1}", fill_color=cur_color)
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=f"""{df["station_name"][i]}""", fill_color=cur_color)
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=f"""{cls._get_platforms(df["station_id"][i])}""", fill_color=cur_color)
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=f"""{df["start_time"][i]}""", fill_color=cur_color)
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=df["total_power"][i], fill_color=cur_color, num_style='#,##0')
               cur_color = "00B050"
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=df["day_charged_count"][i], fill_color=cur_color, num_style='#,##0')
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=df["day_charging_capacity(kwh)"][i], fill_color=cur_color)
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=df["day_utilize_hours"][i], fill_color=cur_color)
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=df["day_electric_fee(RMB)"][i], fill_color=cur_color)
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=df["day_service_fee(RMB)"][i], fill_color=cur_color)
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=df["day_average_service_fee_per_kwh(RMB)"][i], fill_color=cur_color)
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=f"""/""", fill_color=cur_color)
               cur_color = "FFFF00"
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=df["month_charged_count"][i], fill_color=cur_color, num_style='#,##0')
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=df["month_charging_capacity(kwh)"][i], fill_color=cur_color)
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=df["month_utilize_hours"][i], fill_color=cur_color)
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=df["month_electric_fee(RMB)"][i], fill_color=cur_color)
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=df["month_service_fee(RMB)"][i], fill_color=cur_color)
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=f"""/""", fill_color=cur_color)
               cur_color = "00B0F0"
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=df["year_charged_count"][i], fill_color=cur_color, num_style='#,##0')
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=df["year_charging_capacity(kwh)"][i], fill_color=cur_color)
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=df["year_utilize_hours"][i], fill_color=cur_color)
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=df["year_electric_fee(RMB)"][i], fill_color=cur_color)
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=df["year_service_fee(RMB)"][i], fill_color=cur_color)
               cls._set_cell(cell=ws[f"{it.ch}{cur_row}"], value=f"""/""", fill_color=cur_color)
               cur_row += 1
          middle_end = cur_row-1

          """
          倒数第2行，行高27.0：
              字体：11号，加粗
              A-B：内容："合计"; E-W: 背景颜色:"FFC000"     
          """
          it = Iterator()
          ws.row_dimensions[cur_row].height = 27.0
          ch_l = []
          for i in range(4):
               ch_l.append(it.ch)
          ws.merge_cells(f"{ch_l[0]}{cur_row}:{ch_l[-1]}{cur_row}")
          cur_color = "FFFFFF"
          cls._set_cell(cell=ws[f"{ch_l[0]}{cur_row}"], value=f"""合计""", fill_color=cur_color)
          for i in range(4):
               cls._set_cell_border(cell=ws[f"{ch_l[i]}{cur_row}"])

          cur_color = "FFC000"
          ch_l = []
          for i in range(20):
               ch_l.append(it.ch)
          cls._set_cell(cell=ws[f"{ch_l[0]}{cur_row}"], value=f"""=SUM({ch_l[0] + str(middle_start)}:{ch_l[0] + str(middle_end)})""", fill_color="FFFFFF", num_style='#,##0')
          for s in ch_l[2:3]+ch_l[4:6]+ch_l[7:8]+ch_l[9:10]+ch_l[11:14]+ch_l[15:16]+ch_l[17:]:
               cls._set_cell(cell=ws[f"{s}{cur_row}"], value=f"""=SUM({s+str(middle_start)}:{s+str(middle_end)})""", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{ch_l[3]}{cur_row}"], value=f"""={ch_l[2]}{cur_row} / {ch_l[0]}{cur_row}""", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{ch_l[10]}{cur_row}"], value=f"""={ch_l[9]}{cur_row} / {ch_l[0]}{cur_row}""", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{ch_l[16]}{cur_row}"], value=f"""={ch_l[15]}{cur_row} / {ch_l[0]}{cur_row}""", fill_color=cur_color)
          cls._set_cell(cell=ws[f"{ch_l[6]}{cur_row}"], value=f"""={ch_l[5]}{cur_row} / {ch_l[2]}{cur_row}""", fill_color=cur_color)
          for s in [ch_l[1],ch_l[8],ch_l[14]]:
               cls._set_cell(cell=ws[f"{s}{cur_row}"], value=f"""=SUM({s + str(middle_start)}:{s + str(middle_end)})""",
                             fill_color=cur_color, num_style='#,##0')
          cur_row += 1

          # """
          # 最后1行，行高45.0：
          #     字体：11号，不加粗
          #     A-V: 内容："主要事件及工作："
          # """
          # it = Iterator()
          # ch_l = []
          # for i in range(22):
          #      ch_l.append(it.ch)
          # ws.row_dimensions[cur_row].height = 45.0
          # ws.merge_cells(f"{ch_l[0]}{cur_row}:{ch_l[-1]}{cur_row}")
          # cls._set_cell(cell=ws[f"{ch_l[0]}{cur_row}"], value="主要事件及工作：", font_bold=False, horizontal="left")
          # for s in ch_l[1:]:
          #      cls._set_cell_border(ws[f"{s}{cur_row}"])
          # cur_row += 1

          return cur_row, (middle_start, middle_end)

     @classmethod
     def generate_financial_report(cls, year, month, day):
          # reporters = FinancialReportGetter.get_station_order_reporters(year, month, day)
          reporters = FinancialReportGetterAfterStartDate.get_station_order_reporters(year, month, day)
          # 创建一个工作簿（Workbook）
          wb = Workbook()
          # 获取当前活动的工作表（Worksheet）,就是子表
          ws = wb.active
          cls._generate_form(ws)

          cur_date_str = f"{year:04d}-{month:02d}-{day:02d}"
          cur_date_str_dot = cur_date_str.replace("-", ".")

          cur_row = 1
          num_rows = []
          for city, df in reporters.items():
               cur_row, cur_num_rows = cls._generate_city(ws=ws, df=df, city=city, date_str=cur_date_str_dot, cur_row=cur_row)
               num_rows.append(cur_num_rows)

          ws.row_dimensions[cur_row].height = 45.0
          cur_color = "7890F7"

          it = Iterator()
          ch_l = []
          for i in range(4):
               ch_l.append(it.ch)
          ws.merge_cells(f"{ch_l[0]}{cur_row}:{ch_l[-1]}{cur_row}")
          cls._set_cell(cell=ws[f"{ch_l[0]}{cur_row}"], value="全国合计", fill_color=cur_color)
          for s in ch_l[1:]:
               cls._set_cell_border(ws[f"{s}{cur_row}"])

          l_used = []
          for i in range(20):
               s = it.ch
               num_style = '#,##0.00'
               l_used.append(s)
               v = "="
               if i in [3,10,16]:
                    v += f"{l_used[i-1]}{cur_row} / {l_used[0]}{cur_row}"
               elif i == 6:
                    v += f"{l_used[5]}{cur_row} / {l_used[2]}{cur_row}"
               else:
                    for cur_num_rows in num_rows:
                         v += f"SUM({s}{cur_num_rows[0]}:{s}{cur_num_rows[1]})+"
                    v = v[:-1]

                    if i in [0,1,8,14]:
                         num_style = '#,##0'
               cls._set_cell(cell=ws[f"{s}{cur_row}"],
                             value=v,
                             fill_color=cur_color,
                             num_style=num_style)
               cls._set_cell_border(cell=ws[f"{s}{cur_row}"])
          wb.save(f"{cur_date_str}财务报表.xlsx")

     @classmethod
     def generate_txt(cls, year, month, day):
          df_today = FinancialReportGetter.get_station_order_report_day_city(year, month, day).fillna(0)
          previous_day = date(year, month, day) - timedelta(days=1)
          df_pre = FinancialReportGetter.get_station_order_report_day_city(previous_day.year,
                                                                               previous_day.month,
                                                                               previous_day.day).fillna(0)
          dic_today = {}
          for index, row in df_today.iterrows():
               row_dic = row.to_dict()
               dic_today[row_dic["city"]] = row_dic

          dic_pre = {}
          for index, row in df_pre.iterrows():
               row_dic = row.to_dict()
               dic_pre[row_dic["city"]] = row_dic

          str = f"""各位领导，请收阅{month}月{day}日报表，相较于{previous_day.month}月{previous_day.day}日，\n"""
          for k, v in dic_today.items():
               city = k
               v_pre = dic_pre.get(city)
               if v_pre["day_charged_count"] == 0 and v["day_charged_count"] == 0:
                    day_charged_count = 0
               elif v_pre["day_charged_count"] == 0:
                    day_charged_count = 100
               else:
                    day_charged_count = (v["day_charged_count"]/v_pre["day_charged_count"] - 1)*100

               if v_pre["day_charging_capacity(kwh)"] == 0 and v["day_charging_capacity(kwh)"] == 0:
                    day_charging_capacity = 0
               elif v_pre["day_charging_capacity(kwh)"] == 0:
                    day_charging_capacity = 100
               else:
                    day_charging_capacity = (v["day_charging_capacity(kwh)"]/v_pre["day_charging_capacity(kwh)"] - 1)*100

               if v_pre["day_service_fee(RMB)"] == 0 and v["day_service_fee(RMB)"] == 0:
                    service_fee = 0
               elif v_pre["day_service_fee(RMB)"] == 0:
                    service_fee = 100
               else:
                    service_fee = (v["day_service_fee(RMB)"]/v_pre["day_service_fee(RMB)"] - 1)*100

               day_charged_count = round(day_charged_count, 2)
               day_charging_capacity = round(day_charging_capacity, 2)
               service_fee = round(service_fee,2)
               if day_charged_count >= 0:
                    str += f"""【{city}汇天】充电次数上升{abs(day_charged_count)}%，"""
               else:
                    str += f"""【{city}汇天】充电次数下降{abs(day_charged_count)}%，"""

               if day_charging_capacity >= 0:
                    str += f"""充电量上升{abs(day_charging_capacity)}%，"""
               else:
                    str += f"""充电量下降{abs(day_charging_capacity)}%，"""

               if service_fee >= 0:
                    str += f"""服务费上升{abs(service_fee)}%。\n"""
               else:
                    str += f"""服务费下降{abs(service_fee)}%。\n"""

          str = str[:-2]
          str += "，请知悉。"

          cur_date_str = f"{year:04d}-{month:02d}-{day:02d}"
          with open(f"{cur_date_str}环比.txt", 'w', encoding='utf-8') as file:
               file.write(str)

     @classmethod
     def _get_platforms(cls, station):
          l = set(cls.df_station_map[cls.df_station_map["station_id"] == station]["platform"])
          ht = ""
          str = ""
          for s in l:
               if s != "汇天":
                    str += f"{s}-"
               else:
                    ht = "汇天-"
          str = ht+str
          return str[:-1]
